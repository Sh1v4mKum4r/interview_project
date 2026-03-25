"""
Excel report generator — creates a formatted 6-sheet .xlsx workbook
from analysis results using openpyxl.
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11, color="1F4E79")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ROW_ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel_report(
    results: dict,
    returns: pd.DataFrame,
    prices: pd.DataFrame,
    metadata: dict,
    config: dict,
) -> bytes:
    """
    Generate a formatted Excel report with 7 sheets.
    Returns the workbook as bytes.
    """
    wb = Workbook()

    _write_executive_summary(wb, results, metadata, config)
    _write_statistical_analysis(wb, results, metadata)
    _write_risk_model_output(wb, results, config)
    _write_quantitative_analysis(wb, results, metadata)
    _write_advanced_analysis(wb, results)
    _write_regulatory_report(wb, results)
    _write_raw_data(wb, returns, prices, metadata)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Auto-fit column widths for all sheets
    for ws in wb.worksheets:
        _auto_fit_columns(ws)
        ws.sheet_properties.tabColor = "1F4E79"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_subheader_row(ws, row: int, num_cols: int):
    """Apply subheader styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int):
    """Apply alternating row colors and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if (row - start_row) % 2 == 1:
                cell.fill = ROW_ALT_FILL


def _status_fill(status: str) -> PatternFill:
    """Return the appropriate fill for a status string."""
    s = status.lower()
    if s in ("pass", "compliant", "green"):
        return PASS_FILL
    elif s in ("amber", "warning"):
        return AMBER_FILL
    return FAIL_FILL


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)


def _write_executive_summary(wb: Workbook, results: dict, metadata: dict, config: dict):
    """Sheet 1: Executive Summary."""
    ws = wb.create_sheet("Executive Summary", 0)
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Regulatory Risk Analysis — Executive Summary"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    # Portfolio overview
    row = 4
    ws.cell(row=row, column=1, value="Portfolio Overview")
    _style_subheader_row(ws, row, 4)
    row += 1

    overview = [
        ("Total Assets", len(metadata.get("asset_classes", {}))),
        ("Portfolio Value", f"${config.get('portfolio_value', 10_000_000):,.0f}"),
        ("Time Horizon", config.get("time_horizon", "1D")),
        ("Confidence Level", f"{config.get('confidence_level', 0.95)*100:.0f}%"),
    ]
    for label, value in overview:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    _style_data_rows(ws, 5, row - 1, 4)

    # Key risk metrics
    row += 1
    headers = ["Metric", "Value", "Unit"]
    ws.cell(row=row, column=1, value="Key Risk Metrics")
    _style_subheader_row(ws, row, len(headers))
    row += 1

    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    risk = results.get("risk", {}).get("metrics", {})
    hist = risk.get("historical", {})
    metrics_data = [
        ("VaR (Historical)", f"{abs(hist.get('var_dollar', 0)):,.0f}", "USD"),
        ("Expected Shortfall", f"{abs(hist.get('es_dollar', 0)):,.0f}", "USD"),
        ("VaR % of Portfolio", f"{abs(hist.get('portfolio_var', 0))*100:.2f}", "%"),
        ("ES % of Portfolio", f"{abs(hist.get('portfolio_es', 0))*100:.2f}", "%"),
    ]
    start_data = row
    for label, value, unit in metrics_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=unit)
        row += 1
    _style_data_rows(ws, start_data, row - 1, 3)

    # Regulatory compliance
    row += 1
    ws.cell(row=row, column=1, value="Regulatory Compliance")
    _style_subheader_row(ws, row, 4)
    row += 1

    reg_headers = ["Metric", "Value", "Threshold", "Status"]
    for i, h in enumerate(reg_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(reg_headers))
    row += 1

    reg = results.get("regulatory", {}).get("metrics", {})
    basel = reg.get("basel3", {})
    summary_items = [
        ("CET1 Ratio", f"{basel.get('cet1_ratio', 0)*100:.1f}%", "≥ 4.5%", basel.get("cet1_status", "N/A")),
        ("Tier 1 Ratio", f"{basel.get('tier1_ratio', 0)*100:.1f}%", "≥ 6.0%", basel.get("tier1_status", "N/A")),
        ("Total Capital", f"{basel.get('total_capital_ratio', 0)*100:.1f}%", "≥ 8.0%", basel.get("total_capital_status", "N/A")),
        ("Leverage Ratio", f"{basel.get('leverage_ratio', 0)*100:.1f}%", "≥ 3.0%", basel.get("leverage_status", "N/A")),
        ("LCR", f"{basel.get('lcr', 0)*100:.1f}%", "≥ 100%", basel.get("lcr_status", "N/A")),
    ]
    start_data = row
    for label, value, threshold, status in summary_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=threshold)
        status_cell = ws.cell(row=row, column=4, value=status)
        status_cell.fill = _status_fill(status)
        row += 1
    _style_data_rows(ws, start_data, row - 1, 4)


def _write_statistical_analysis(wb: Workbook, results: dict, metadata: dict):
    """Sheet 2: Statistical Analysis."""
    ws = wb.create_sheet("Statistical Analysis")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Statistical Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    row = 3

    # Moments table
    ws.cell(row=row, column=1, value="Distribution Moments (Annualized)")
    _style_subheader_row(ws, row, 5)
    row += 1

    headers = ["Asset", "Mean", "Variance", "Skewness", "Kurtosis"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    moments = results.get("moments", {}).get("metrics", {})
    start_data = row
    for asset, vals in moments.items():
        if isinstance(vals, dict):
            ws.cell(row=row, column=1, value=asset)
            ws.cell(row=row, column=2, value=round(vals.get("mean", 0), 6))
            ws.cell(row=row, column=3, value=round(vals.get("variance", 0), 6))
            ws.cell(row=row, column=4, value=round(vals.get("skewness", 0), 4))
            ws.cell(row=row, column=5, value=round(vals.get("kurtosis", 0), 4))
            row += 1
    _style_data_rows(ws, start_data, row - 1, 5)

    # Correlation matrix
    row += 1
    ws.cell(row=row, column=1, value="Correlation Matrix")
    _style_subheader_row(ws, row, 5)
    row += 1

    corr_data = results.get("correlation", {}).get("data", {})
    labels = corr_data.get("labels", [])
    matrix = corr_data.get("matrix", [])

    if labels:
        # Header row with asset names
        for i, label in enumerate(labels):
            ws.cell(row=row, column=i + 2, value=label)
        _style_header_row(ws, row, len(labels) + 1)
        row += 1

        start_data = row
        for i, label in enumerate(labels):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            for j, val in enumerate(matrix[i] if i < len(matrix) else []):
                cell = ws.cell(row=row, column=j + 2, value=round(val, 3))
                cell.number_format = "0.000"
            row += 1
        _style_data_rows(ws, start_data, row - 1, len(labels) + 1)

    # Distribution fitting
    row += 1
    ws.cell(row=row, column=1, value="Distribution Fitting (Best Fit)")
    _style_subheader_row(ws, row, 4)
    row += 1

    fit_headers = ["Asset", "Best Fit Distribution", "AIC (Best)", "Parameters"]
    for i, h in enumerate(fit_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(fit_headers))
    row += 1

    dist_fit = results.get("distribution_fitting", {}).get("metrics", {})
    start_data = row
    for asset, info in dist_fit.items():
        if isinstance(info, dict):
            ws.cell(row=row, column=1, value=asset)
            ws.cell(row=row, column=2, value=info.get("best_fit", "N/A"))
            aic_vals = info.get("aic", {})
            best = info.get("best_fit", "")
            ws.cell(row=row, column=3, value=round(aic_vals.get(best, 0), 2) if isinstance(aic_vals, dict) else "N/A")
            params = info.get("params", {})
            ws.cell(row=row, column=4, value=str(params.get(best, {})) if isinstance(params, dict) else "N/A")
            row += 1
    _style_data_rows(ws, start_data, row - 1, 4)


def _write_risk_model_output(wb: Workbook, results: dict, config: dict):
    """Sheet 3: Risk Model Output."""
    ws = wb.create_sheet("Risk Model Output")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Risk Model Output"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    row = 3

    # VaR comparison
    ws.cell(row=row, column=1, value="VaR Model Comparison")
    _style_subheader_row(ws, row, 5)
    row += 1

    headers = ["Model", "VaR (%)", "VaR ($)", "ES (%)", "ES ($)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    risk = results.get("risk", {}).get("metrics", {})
    pv = config.get("portfolio_value", 10_000_000)
    models = [
        ("Historical", risk.get("historical", {})),
        ("Parametric (Normal)", risk.get("parametric_normal", {})),
        ("Parametric (Student-t)", risk.get("parametric_t", {})),
        ("Monte Carlo", risk.get("monte_carlo", {})),
    ]

    start_data = row
    for name, data in models:
        ws.cell(row=row, column=1, value=name)
        var_pct = data.get("portfolio_var", 0)
        es_pct = data.get("portfolio_es", 0)
        ws.cell(row=row, column=2, value=f"{abs(var_pct)*100:.2f}%")
        ws.cell(row=row, column=3, value=f"${abs(data.get('var_dollar', var_pct * pv)):,.0f}")
        ws.cell(row=row, column=4, value=f"{abs(es_pct)*100:.2f}%")
        ws.cell(row=row, column=5, value=f"${abs(data.get('es_dollar', es_pct * pv)):,.0f}")
        row += 1
    _style_data_rows(ws, start_data, row - 1, 5)

    # Per-asset VaR (historical)
    row += 1
    ws.cell(row=row, column=1, value="Historical VaR by Asset")
    _style_subheader_row(ws, row, 3)
    row += 1

    asset_headers = ["Asset", "VaR (%)", "ES (%)"]
    for i, h in enumerate(asset_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(asset_headers))
    row += 1

    per_asset = risk.get("historical", {}).get("per_asset", {})
    start_data = row
    for asset, vals in per_asset.items():
        if isinstance(vals, dict):
            ws.cell(row=row, column=1, value=asset)
            ws.cell(row=row, column=2, value=f"{abs(vals.get('var', 0))*100:.2f}%")
            ws.cell(row=row, column=3, value=f"{abs(vals.get('es', 0))*100:.2f}%")
            row += 1
    _style_data_rows(ws, start_data, row - 1, 3)


def _write_quantitative_analysis(wb: Workbook, results: dict, metadata: dict):
    """Sheet 4: Quantitative Analysis."""
    ws = wb.create_sheet("Quantitative Analysis")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Quantitative Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    row = 3

    # PCA results
    ws.cell(row=row, column=1, value="PCA — Variance Explained")
    _style_subheader_row(ws, row, 4)
    row += 1

    pca = results.get("pca", {}).get("metrics", {})
    pca_headers = ["Component", "Eigenvalue", "Variance Explained", "Cumulative"]
    for i, h in enumerate(pca_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(pca_headers))
    row += 1

    eigenvalues = pca.get("eigenvalues", [])
    var_explained = pca.get("variance_explained", [])
    cum_var = pca.get("cumulative_variance", [])
    start_data = row
    for i in range(len(eigenvalues)):
        ws.cell(row=row, column=1, value=f"PC{i+1}")
        ws.cell(row=row, column=2, value=round(eigenvalues[i], 4) if i < len(eigenvalues) else "")
        ws.cell(row=row, column=3, value=f"{var_explained[i]*100:.1f}%" if i < len(var_explained) else "")
        ws.cell(row=row, column=4, value=f"{cum_var[i]*100:.1f}%" if i < len(cum_var) else "")
        row += 1
    _style_data_rows(ws, start_data, row - 1, 4)

    # Cluster assignments
    row += 1
    clustering = results.get("clustering", {}).get("metrics", {})
    ws.cell(row=row, column=1, value=f"K-Means Clustering (k={clustering.get('n_clusters', 'N/A')})")
    _style_subheader_row(ws, row, 2)
    row += 1

    cl_headers = ["Asset", "Cluster"]
    for i, h in enumerate(cl_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(cl_headers))
    row += 1

    assignments = clustering.get("assignments", {})
    start_data = row
    for asset, cluster in assignments.items():
        ws.cell(row=row, column=1, value=asset)
        ws.cell(row=row, column=2, value=cluster)
        row += 1
    _style_data_rows(ws, start_data, row - 1, 2)

    # Regression
    row += 1
    ws.cell(row=row, column=1, value="Multi-Factor Regression (Fama-French)")
    _style_subheader_row(ws, row, 6)
    row += 1

    reg_headers = ["Asset", "Alpha", "Beta (MKT)", "Beta (SMB)", "Beta (HML)", "R²"]
    for i, h in enumerate(reg_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(reg_headers))
    row += 1

    regression = results.get("regression", {}).get("metrics", {})
    start_data = row
    for asset, vals in regression.items():
        if isinstance(vals, dict):
            ws.cell(row=row, column=1, value=asset)
            ws.cell(row=row, column=2, value=round(vals.get("alpha", 0), 6))
            ws.cell(row=row, column=3, value=round(vals.get("beta_mkt", 0), 4))
            ws.cell(row=row, column=4, value=round(vals.get("beta_smb", 0), 4))
            ws.cell(row=row, column=5, value=round(vals.get("beta_hml", 0), 4))
            ws.cell(row=row, column=6, value=round(vals.get("r_squared", 0), 4))
            row += 1
    _style_data_rows(ws, start_data, row - 1, 6)

    # Exposure
    row += 1
    ws.cell(row=row, column=1, value="Asset Class Exposure")
    _style_subheader_row(ws, row, 3)
    row += 1

    exp_headers = ["Asset Class", "Weight", "Risk Contribution"]
    for i, h in enumerate(exp_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(exp_headers))
    row += 1

    exposure = results.get("exposure", {}).get("metrics", {})
    weight_by_class = exposure.get("weight_by_class", {})
    rc_by_class = exposure.get("risk_contribution_by_class", {})
    start_data = row
    for cls in weight_by_class:
        ws.cell(row=row, column=1, value=cls)
        ws.cell(row=row, column=2, value=f"{weight_by_class.get(cls, 0)*100:.1f}%")
        ws.cell(row=row, column=3, value=f"{rc_by_class.get(cls, 0)*100:.1f}%")
        row += 1
    _style_data_rows(ws, start_data, row - 1, 3)


def _write_advanced_analysis(wb: Workbook, results: dict):
    """Sheet 5: Advanced Quantitative Analysis."""
    ws = wb.create_sheet("Advanced Analysis")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Advanced Quantitative Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    row = 3
    advanced = results.get("advanced", {}).get("metrics", {})

    # Taylor Series
    ws.cell(row=row, column=1, value="Taylor Series (Delta-Gamma) Approximation")
    _style_subheader_row(ws, row, 3)
    row += 1

    taylor = advanced.get("taylor_series", {})
    taylor_items = [
        ("Delta-only VaR", taylor.get("delta_only_var", 0)),
        ("Delta-Gamma VaR", taylor.get("delta_gamma_var", 0)),
        ("Gamma Correction", taylor.get("gamma_correction", 0)),
    ]
    start_data = row
    for label, val in taylor_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(val, 2))
        row += 1
    _style_data_rows(ws, start_data, row - 1, 2)

    # Laplace Transforms
    row += 1
    ws.cell(row=row, column=1, value="Laplace Transforms — Aggregate Loss Modelling")
    _style_subheader_row(ws, row, 3)
    row += 1

    laplace = advanced.get("laplace_transforms", {})
    laplace_items = [
        ("Expected Loss (E[S])", laplace.get("expected_loss", 0)),
        ("Aggregate VaR (95%)", laplace.get("aggregate_var_95", 0)),
        ("Aggregate VaR (99%)", laplace.get("aggregate_var_99", 0)),
        ("Aggregate ES (95%)", laplace.get("aggregate_es_95", 0)),
        ("Aggregate ES (99%)", laplace.get("aggregate_es_99", 0)),
    ]
    start_data = row
    for label, val in laplace_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(val, 2))
        row += 1
    _style_data_rows(ws, start_data, row - 1, 2)

    # EVT / GPD
    row += 1
    ws.cell(row=row, column=1, value="Extreme Value Theory (Peaks-Over-Threshold / GPD)")
    _style_subheader_row(ws, row, 3)
    row += 1

    evt = advanced.get("evt_gpd", {})
    evt_items = [
        ("GPD Shape Parameter (xi)", evt.get("gpd_shape_xi", 0)),
        ("GPD Scale Parameter (beta)", evt.get("gpd_scale_beta", 0)),
        ("Threshold (u)", evt.get("threshold", 0)),
        ("Number of Exceedances", evt.get("n_exceedances", 0)),
        ("Tail VaR (99%)", evt.get("tail_var_99", 0)),
        ("Tail ES (99%)", evt.get("tail_es_99", 0)),
    ]
    start_data = row
    for label, val in evt_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(val, 4))
        row += 1
    _style_data_rows(ws, start_data, row - 1, 2)


def _write_regulatory_report(wb: Workbook, results: dict):
    """Sheet 5: Regulatory Report."""
    ws = wb.create_sheet("Regulatory Report")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Regulatory Report"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    row = 3
    reg = results.get("regulatory", {}).get("metrics", {})
    reg_data = results.get("regulatory", {}).get("data", {})

    # Basel III/IV
    ws.cell(row=row, column=1, value="Basel III/IV Capital Adequacy")
    _style_subheader_row(ws, row, 4)
    row += 1

    headers = ["Metric", "Value", "Minimum Threshold", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    basel = reg.get("basel3", {})
    summary_table = reg_data.get("summary_table", reg_data.get("basel3", {}).get("capital_ratios", {}))

    basel_items = [
        ("CET1 Ratio", f"{basel.get('cet1_ratio', 0)*100:.2f}%", "≥ 4.5%", basel.get("cet1_status", "N/A")),
        ("Tier 1 Ratio", f"{basel.get('tier1_ratio', 0)*100:.2f}%", "≥ 6.0%", basel.get("tier1_status", "N/A")),
        ("Total Capital Ratio", f"{basel.get('total_capital_ratio', 0)*100:.2f}%", "≥ 8.0%", basel.get("total_capital_status", "N/A")),
        ("Leverage Ratio", f"{basel.get('leverage_ratio', 0)*100:.2f}%", "≥ 3.0%", basel.get("leverage_status", "N/A")),
        ("LCR", f"{basel.get('lcr', 0)*100:.1f}%", "≥ 100%", basel.get("lcr_status", "N/A")),
        ("FRTB ES (97.5%)", f"${abs(basel.get('frtb_es', 0)):,.0f}", "N/A", "info"),
    ]

    start_data = row
    for label, value, threshold, status in basel_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=threshold)
        status_cell = ws.cell(row=row, column=4, value=status)
        if status != "info":
            status_cell.fill = _status_fill(status)
        row += 1
    _style_data_rows(ws, start_data, row - 1, 4)

    # RWA breakdown
    row += 1
    ws.cell(row=row, column=1, value="Risk-Weighted Assets Breakdown")
    _style_subheader_row(ws, row, 2)
    row += 1

    rwa_headers = ["Asset Class", "RWA"]
    for i, h in enumerate(rwa_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(rwa_headers))
    row += 1

    rwa_breakdown = basel.get("rwa_breakdown", {})
    start_data = row
    for cls, rwa in rwa_breakdown.items():
        ws.cell(row=row, column=1, value=cls)
        ws.cell(row=row, column=2, value=f"${rwa:,.0f}")
        row += 1
    ws.cell(row=row, column=1, value="Total RWA")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"${basel.get('rwa', 0):,.0f}")
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    _style_data_rows(ws, start_data, row - 1, 2)

    # MiFID II
    row += 1
    mifid = reg.get("mifid2", {})
    ws.cell(row=row, column=1, value="MiFID II Compliance Summary")
    _style_subheader_row(ws, row, 2)
    row += 1

    mifid_items = [
        ("Transaction Reports Generated", mifid.get("transaction_count", 0)),
        ("Best Execution Flags", mifid.get("best_execution_flags", 0)),
        ("Position Limit Breaches", len(mifid.get("position_limit_breaches", []))),
    ]
    start_data = row
    for label, value in mifid_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _style_data_rows(ws, start_data, row - 1, 2)


def _write_raw_data(wb: Workbook, returns: pd.DataFrame, prices: pd.DataFrame, metadata: dict):
    """Sheet 6: Raw Data."""
    ws = wb.create_sheet("Raw Data")
    ws.freeze_panes = "B2"

    # Returns
    ws.cell(row=1, column=1, value="Daily Log-Returns")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    # Header row
    ws.cell(row=2, column=1, value="Date")
    for j, col in enumerate(returns.columns):
        ws.cell(row=2, column=j + 2, value=col)
    _style_header_row(ws, 2, len(returns.columns) + 1)

    # Data rows (limit to 300 rows for file size)
    max_rows = min(len(returns), 300)
    for i in range(max_rows):
        ws.cell(row=i + 3, column=1, value=str(returns.index[i].date()))
        for j, col in enumerate(returns.columns):
            val = returns.iloc[i, j]
            cell = ws.cell(row=i + 3, column=j + 2, value=round(float(val), 6))
            cell.number_format = "0.000000"

    # Asset metadata at the bottom
    meta_start = max_rows + 5
    ws.cell(row=meta_start, column=1, value="Asset Metadata")
    ws.cell(row=meta_start, column=1).font = Font(bold=True, size=12, color="1F4E79")

    meta_headers = ["Asset", "Class", "Sector", "Rating"]
    for i, h in enumerate(meta_headers, 1):
        ws.cell(row=meta_start + 1, column=i, value=h)
    _style_header_row(ws, meta_start + 1, len(meta_headers))

    ac = metadata.get("asset_classes", {})
    sec = metadata.get("sectors", {})
    rat = metadata.get("ratings", {})

    for idx, asset in enumerate(ac.keys()):
        r = meta_start + 2 + idx
        ws.cell(row=r, column=1, value=asset)
        ws.cell(row=r, column=2, value=ac.get(asset, ""))
        ws.cell(row=r, column=3, value=sec.get(asset, ""))
        ws.cell(row=r, column=4, value=rat.get(asset, ""))
